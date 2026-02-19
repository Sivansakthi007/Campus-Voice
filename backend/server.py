from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import JSONResponse
from fastapi.exceptions import RequestValidationError
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from sqlalchemy import select, func, text, or_
from sqlalchemy.ext.asyncio import AsyncSession
from db import get_engine, get_session, Base
from models import User, Complaint, StaffRating, HODRating, HODReportToggle
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
from enum import Enum
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
import base64
import io
from utils.rbac import require_roles, is_valid_transition
from functools import wraps

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# SQLAlchemy (MySQL) engine/session
SQLALCHEMY_DATABASE_URL = os.environ.get('SQLALCHEMY_DATABASE_URL')

# Security
pwd_context = CryptContext(schemes=["pbkdf2_sha256", "bcrypt"], deprecated="auto")
security = HTTPBearer()
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 1 week

# AI Configuration
LLM_KEY = os.environ.get('LLM_KEY')

# Create the main app
app = FastAPI(title="Campus Voice API")
api_router = APIRouter(prefix="/api")

# Add CORS middleware - Production Ready
ALLOWED_ORIGINS = [
    "https://campus-voice-frontend.onrender.com",  # Production frontend
    "http://localhost:3000",  # Local development
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["Content-Type", "Authorization", "Accept", "Origin", "X-Requested-With"],
)

@app.on_event("startup")
async def startup():
    """Initialize database connection with retry logic for containerized environments."""
    import asyncio
    
    max_retries = 5
    retry_delay = 2  # Start with 2 seconds
    
    for attempt in range(1, max_retries + 1):
        try:
            logger.info(f"Attempting database connection (attempt {attempt}/{max_retries})...")
            engine = get_engine()
            async with engine.begin() as conn:
                await conn.run_sync(Base.metadata.create_all)
            
            # Safe migration: add staff_role column if it doesn't exist
            try:
                async with engine.begin() as conn:
                    # Check if staff_role column exists
                    result = await conn.execute(text(
                        "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                        "WHERE TABLE_NAME = 'users' AND COLUMN_NAME = 'staff_role'"
                    ))
                    column_exists = result.fetchone() is not None
                    
                    if not column_exists:
                        await conn.execute(text(
                            "ALTER TABLE users ADD COLUMN staff_role VARCHAR(100) NULL"
                        ))
                        logger.info("Migration: Added 'staff_role' column to users table")
                    else:
                        logger.info("Migration: 'staff_role' column already exists")
            except Exception as migration_err:
                logger.warning(f"Migration check for staff_role column: {str(migration_err)}")
            
            # Safe migration: add student_department column to complaints if it doesn't exist
            try:
                async with engine.begin() as conn:
                    result = await conn.execute(text(
                        "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                        "WHERE TABLE_NAME = 'complaints' AND COLUMN_NAME = 'student_department'"
                    ))
                    column_exists = result.fetchone() is not None
                    
                    if not column_exists:
                        await conn.execute(text(
                            "ALTER TABLE complaints ADD COLUMN student_department VARCHAR(255) NULL"
                        ))
                        logger.info("Migration: Added 'student_department' column to complaints table")
                    else:
                        logger.info("Migration: 'student_department' column already exists")
            except Exception as migration_err:
                logger.warning(f"Migration check for student_department column: {str(migration_err)}")
            
            logger.info("Database connection established successfully!")
            return
        except Exception as e:
            logger.warning(f"Database connection attempt {attempt} failed: {str(e)}")
            if attempt < max_retries:
                wait_time = min(retry_delay * (2 ** (attempt - 1)), 32)  # Exponential backoff, max 32s
                logger.info(f"Retrying in {wait_time} seconds...")
                await asyncio.sleep(wait_time)
            else:
                logger.error("All database connection attempts failed!")
                raise
@app.get("/")
async def root():
    return {"message": "Campus Voice API is running", "docs_url": "/docs"}

# Health check endpoint
@api_router.get("/health")
async def health_check():
    """Health check endpoint for monitoring and load balancers"""
    return {"status": "healthy", "service": "Campus Voice API"}

# ===== MODELS =====
class UserRole(str):
    STUDENT = "student"
    ADMIN = "admin"
    HOD = "hod"
    PRINCIPAL = "principal"
    STAFF = "staff"

class ComplaintStatus(str):
    SUBMITTED = "submitted"
    REVIEWED = "reviewed"
    IN_PROGRESS = "in_progress"
    RESOLVED = "resolved"
    REJECTED = "rejected"

class ComplaintCategory(str):
    SPORTS = "Sports"
    LIBRARY = "Library"
    DISCIPLINE = "Discipline"
    EXAM_CELL = "Exam Cell"
    ACCOUNTS_FEES = "Accounts/Fees"
    TRANSPORT = "Transport"
    SCHOLARSHIP = "Scholarship"
    PLACEMENT_TRAINING = "Placement/Training"
    HOSTEL = "Hostel"
    INFRASTRUCTURE = "Infrastructure"
    LAB = "Lab"
    ACADEMIC = "Academic Issues"
    STAFF_BEHAVIOR = "Staff Behavior"
    OTHER = "Other"

# Staff Role options for registration
class StaffRole(str):
    ASSISTANT_PROFESSOR = "Assistant Professor"
    LAB_ASSISTANT = "Lab Assistant"
    LIBRARIAN = "Librarian"
    PHYSICAL_DIRECTOR = "Physical Director"
    DISCIPLINE_COORDINATOR = "Discipline Coordinator"
    EXAM_CELL_COORDINATOR = "Exam Cell Coordinator"
    ACCOUNTANT = "Accountant"
    CLERK = "Clerk"
    TRANSPORT_MANAGER = "Transport Manager"
    SCHOLARSHIP_COORDINATOR = "Scholarship Coordinator"
    PLACEMENT_TRAINING_COORDINATOR = "Placement & Training Coordinator"
    WARDEN = "Warden"
    INFRASTRUCTURE_COORDINATOR = "Infrastructure Coordinator"

# Categories that MUST be assigned ONLY to the HOD of the student's department
# Manual override is blocked for these categories
HOD_CATEGORIES = {"Academic Issues", "Staff Behavior"}

# Mapping: Complaint Category -> Staff Role for auto-assignment
# NOTE: HOD_CATEGORIES are NOT in this mapping — they use department-based HOD routing
CATEGORY_TO_STAFF_ROLE = {
    "Sports": "Physical Director",
    "Library": "Librarian",
    "Discipline": "Discipline Coordinator",
    "Exam Cell": "Exam Cell Coordinator",
    "Accounts/Fees": "Accountant",
    "Transport": "Transport Manager",
    "Scholarship": "Scholarship Coordinator",
    "Placement/Training": "Placement & Training Coordinator",
    "Hostel": "Warden",
    "Infrastructure": "Infrastructure Coordinator",
    "Lab": "Lab Assistant",
    "Office": "Clerk",
}

class SentimentType(str):
    POSITIVE = "Positive"
    NEGATIVE = "Negative"
    ANGRY = "Angry"
    URGENT = "Urgent"

class PriorityLevel(str):
    HIGH = "High"
    MEDIUM = "Medium"
    LOW = "Low"

class FoulLanguageSeverity(str):
    NONE = "None"
    MILD = "Mild"
    MODERATE = "Moderate"
    SEVERE = "Severe"

# User Models
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str
    department: Optional[str] = None
    student_id: Optional[str] = None
    staff_id: Optional[str] = None
    staff_role: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    role: str
    department: Optional[str] = None
    student_id: Optional[str] = None
    staff_id: Optional[str] = None
    staff_role: Optional[str] = None
    created_at: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

# Complaint Models
class AIAnalysis(BaseModel):
    sentiment: str
    category: str
    priority: str
    foul_language_severity: str
    foul_language_detected: bool

class ComplaintCreate(BaseModel):
    title: str
    description: str
    is_anonymous: bool = False
    voice_text: Optional[str] = None
    category: Optional[str] = None

class ComplaintResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    title: str
    description: str
    status: str
    category: str
    priority: str
    sentiment: str
    foul_language_severity: str
    foul_language_detected: bool
    is_anonymous: bool
    student_id: str
    student_name: Optional[str]
    student_email: Optional[str] = None
    support_count: int
    assigned_to: Optional[str] = None
    assigned_to_name: Optional[str] = None
    assigned_at: Optional[str] = None
    created_at: str
    updated_at: str
    responses: List[Dict[str, Any]]
    timeline: List[Dict[str, Any]]

class ComplaintUpdate(BaseModel):
    status: Optional[str] = None
    response_text: Optional[str] = None
    assigned_to: Optional[str] = None

class SupportVote(BaseModel):
    complaint_id: str

# Analytics Models
class StaffPerformance(BaseModel):
    staff_id: str
    staff_name: str
    total_complaints: int
    resolved_complaints: int
    pending_complaints: int
    avg_resolution_time: float

# Staff Rating Models
class StaffRatingCreate(BaseModel):
    """Request model for submitting a staff rating."""
    staff_id: str
    subject_knowledge: int = Field(..., ge=1, le=5)
    teaching_clarity: int = Field(..., ge=1, le=5)
    student_interaction: int = Field(..., ge=1, le=5)
    punctuality: int = Field(..., ge=1, le=5)
    overall_effectiveness: int = Field(..., ge=1, le=5)

class StaffRatingResponse(BaseModel):
    """Response model for a staff rating."""
    id: str
    staff_id: str
    staff_name: str
    week_number: int
    year: int
    subject_knowledge: int
    teaching_clarity: int
    student_interaction: int
    punctuality: int
    overall_effectiveness: int
    average_rating: float
    created_at: str

class StaffPerformanceRating(BaseModel):
    """Staff performance data for HOD weekly report."""
    staff_id: str
    staff_name: str
    department: Optional[str]
    average_rating: float
    total_ratings: int
    is_best_staff: bool = False

class WeeklyPerformanceReport(BaseModel):
    """Weekly staff performance report for HOD."""
    week_number: int
    year: int
    week_start: str
    week_end: str
    staff_performance: List[StaffPerformanceRating]
    total_ratings: int

# ===== HELPER FUNCTIONS =====
def create_response(success: bool, message: str, data: Any = None, status_code: int = 200) -> JSONResponse:
    """Create a consistent API response format"""
    return JSONResponse(
        status_code=status_code,
        content={
            "success": success,
            "message": message,
            "data": data
        }
    )

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security), session: AsyncSession = Depends(get_session)) -> dict:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication")
        
        user = await session.get(User, user_id)
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        # Compute student_id/staff_id based on role
        stored_id = user.student_id
        is_student = user.role == UserRole.STUDENT
        return {
            "id": user.id,
            "email": user.email,
            "name": user.name,
            "role": user.role,
            "department": user.department,
            "student_id": stored_id if is_student else None,
            "staff_id": stored_id if not is_student else None,
            "staff_role": user.staff_role,
            "created_at": user.created_at.isoformat() if user.created_at else None
        }
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception as e:
        raise HTTPException(status_code=401, detail="Invalid authentication")

def require_roles(*allowed_roles: str):
    def role_checker(current_user: dict = Depends(get_current_user)):
        user = current_user()
        if not user or user.get("role") not in allowed_roles:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have enough permissions to access this resource"
            )
        return current_user
    return role_checker

async def analyze_complaint_with_ai(text: str) -> AIAnalysis:
    """Analyze complaint for sentiment, category, priority, and foul language using OpenAI."""
    try:
        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=LLM_KEY)

        system_message = """You are an AI assistant that analyzes student complaints. 
            Analyze the complaint and respond ONLY with a JSON object (no markdown, no explanation) with these exact keys:
            - sentiment: one of [Positive, Negative, Angry, Urgent]
            - category: one of [Sports, Library, Discipline, Exam Cell, Accounts/Fees, Transport, Scholarship, Placement/Training, Hostel, Infrastructure, Lab, Academic Issues, Staff Behavior, Other]
            - priority: one of [High, Medium, Low]
            - foul_language_detected: boolean (true if foul/offensive language is present)
            - foul_language_severity: one of [None, Mild, Moderate, Severe]

            Consider urgency, emotional tone, and severity when assigning priority."""

        user_prompt = f"Analyze this complaint: {text}"

        # Use chat completions; handle multiple response shapes for resilience
        resp = await client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "system", "content": system_message}, {"role": "user", "content": user_prompt}],
            max_tokens=500
        )

        # Extract text from various possible response shapes
        content = None
        try:
            content = resp.choices[0].message.content
        except Exception:
            try:
                content = resp.choices[0].text
            except Exception:
                content = getattr(resp, "output_text", None) or str(resp)

        import json
        analysis_data = json.loads(content.strip())

        return AIAnalysis(
            sentiment=analysis_data["sentiment"],
            category=analysis_data["category"],
            priority=analysis_data["priority"],
            foul_language_detected=analysis_data["foul_language_detected"],
            foul_language_severity=analysis_data["foul_language_severity"]
        )
    except Exception as e:
        logger.error(f"AI analysis error: {str(e)}")
        # Fallback to default values
        return AIAnalysis(
            sentiment=SentimentType.NEGATIVE,
            category=ComplaintCategory.ACADEMIC,
            priority=PriorityLevel.MEDIUM,
            foul_language_detected=False,
            foul_language_severity=FoulLanguageSeverity.NONE
        )

async def transcribe_audio_with_whisper(audio_base64: str) -> str:
    """Use OpenAI Whisper to transcribe audio to text"""
    try:
        from openai import AsyncOpenAI
        
        client = AsyncOpenAI(api_key=LLM_KEY)
        
        # Decode base64 audio
        audio_bytes = base64.b64decode(audio_base64)
        audio_file = io.BytesIO(audio_bytes)
        audio_file.name = "audio.webm"
        
        # Transcribe using Whisper
        transcription = await client.audio.transcriptions.create(
            model="whisper-1",
            file=audio_file
        )
        
        return transcription.text
    except Exception as e:
        logger.error(f"Whisper transcription error: {str(e)}")
        raise HTTPException(status_code=500, detail="Audio transcription failed")

async def check_duplicate_complaint(title: str, description: str, user_id: str, session: AsyncSession) -> Optional[str]:
    """Check if similar complaint exists in last 30 days"""
    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)

    stmt = select(Complaint.id, Complaint.title, Complaint.description).where(
        Complaint.student_id == user_id,
        Complaint.created_at >= thirty_days_ago
    )
    result = await session.execute(stmt)
    rows = result.all()

    title_lower = title.lower()
    desc_lower = description.lower()

    for row in rows:
        comp_id, comp_title, comp_desc = row[0], row[1], row[2]
        # Check if both title and description are very similar to avoid false positives
        title_match = comp_title and (title_lower in comp_title.lower() or comp_title.lower() in title_lower)
        desc_match = comp_desc and (desc_lower in comp_desc.lower() or comp_desc.lower() in desc_lower)
        
        if title_match and desc_match:
            return comp_id

    return None

# ===== AUTH ENDPOINTS =====
@api_router.post("/auth/register")
async def register(user_data: UserCreate, session: AsyncSession = Depends(get_session)):
    # Check if user exists
    stmt = select(User).where(User.email == user_data.email)
    res = await session.execute(stmt)
    existing_user = res.scalars().first()
    if existing_user:
        return create_response(False, "Email already registered", status_code=400)

    # Validate staff_role for staff registrations
    VALID_STAFF_ROLES = [
        "Assistant Professor", "Lab Assistant", "Librarian",
        "Physical Director", "Discipline Coordinator", "Exam Cell Coordinator",
        "Accountant", "Clerk", "Transport Manager",
        "Scholarship Coordinator", "Placement & Training Coordinator",
        "Warden", "Infrastructure Coordinator"
    ]
    if user_data.role == UserRole.STAFF:
        if not user_data.staff_role or not user_data.staff_role.strip():
            return create_response(False, "Staff Role is required for staff registration. Please select a valid Staff Role.", status_code=400)
        if user_data.staff_role not in VALID_STAFF_ROLES:
            return create_response(False, f"Invalid Staff Role: '{user_data.staff_role}'. Please select a valid Staff Role from the dropdown.", status_code=400)

    # Create user - store ID in student_id column based on role
    is_student = user_data.role == UserRole.STUDENT
    stored_id = user_data.student_id if is_student else user_data.staff_id
    try:
        user_obj = User(
            email=user_data.email,
            password=hash_password(user_data.password),
            name=user_data.name,
            role=user_data.role,
            department=user_data.department,
            student_id=stored_id,  # Polymorphic: stores student_id or staff_id
            staff_role=user_data.staff_role if user_data.role != UserRole.STUDENT else None
        )
        session.add(user_obj)
        await session.commit()
        await session.refresh(user_obj)
    except Exception as e:
        logger.error(f"Registration failed for {user_data.email}: {str(e)}")
        await session.rollback()
        return create_response(False, "Registration failed due to a server error. Please try again.", status_code=500)

    # Create token
    access_token = create_access_token({"sub": user_obj.id})

    # Return computed student_id/staff_id based on role
    stored_id = user_obj.student_id
    is_student_role = user_obj.role == UserRole.STUDENT
    user_response = UserResponse(
        id=user_obj.id,
        email=user_obj.email,
        name=user_obj.name,
        role=user_obj.role,
        department=user_obj.department,
        student_id=stored_id if is_student_role else None,
        staff_id=stored_id if not is_student_role else None,
        staff_role=user_obj.staff_role,
        created_at=user_obj.created_at.isoformat() if user_obj.created_at else None
    )

    data = {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user_response.model_dump()
    }
    return create_response(True, "Registration successful", data)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin, session: AsyncSession = Depends(get_session)):
    stmt = select(User).where(User.email == credentials.email)
    res = await session.execute(stmt)
    user_obj = res.scalars().first()
    if not user_obj or not verify_password(credentials.password, user_obj.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    access_token = create_access_token({"sub": user_obj.id})

    # Return computed student_id/staff_id based on role
    stored_id = user_obj.student_id
    is_student_role = user_obj.role == UserRole.STUDENT
    user_response = UserResponse(
        id=user_obj.id,
        email=user_obj.email,
        name=user_obj.name,
        role=user_obj.role,
        department=user_obj.department,
        student_id=stored_id if is_student_role else None,
        staff_id=stored_id if not is_student_role else None,
        staff_role=user_obj.staff_role,
        created_at=user_obj.created_at.isoformat() if user_obj.created_at else None
    )

    data = {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user_response.model_dump()
    }
    return create_response(True, "Login successful", data)

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return create_response(True, "Profile retrieved successfully", current_user)


@api_router.get("/users")
async def list_users(role: Optional[str] = None, current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    # Only authenticated users can list users; restrict sensitive info
    stmt = select(User)
    if role:
        stmt = stmt.where(User.role == role)

    res = await session.execute(stmt)
    users = res.scalars().all()

    def _to_resp(u: User):
        stored_id = u.student_id
        is_student = u.role == UserRole.STUDENT
        return {
            "id": u.id,
            "email": u.email,
            "name": u.name,
            "role": u.role,
            "department": u.department,
            "student_id": stored_id if is_student else None,
            "staff_id": stored_id if not is_student else None,
            "staff_role": u.staff_role,
            "created_at": u.created_at.isoformat() if u.created_at else None,
        }

    return create_response(True, "Users retrieved successfully", [_to_resp(u) for u in users])

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    """
    Delete a user (Admin only).
    """
    # Only admin can delete users
    if current_user["role"] != UserRole.ADMIN:
        return create_response(False, "Permission denied", status_code=403)
    
    # Prevent admin from deleting themselves
    if current_user["id"] == user_id:
        return create_response(False, "Cannot delete yourself", status_code=400)
    
    user_obj = await session.get(User, user_id)
    if not user_obj:
        return create_response(False, "User not found", status_code=404)
    
    await session.delete(user_obj)
    await session.commit()
    logger.info(f"User {user_id} deleted by Admin {current_user['id']}")
    
    return create_response(True, "User deleted successfully")



@api_router.post("/auth/upload-profile-photo")
async def upload_profile_photo(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    # Validate file type
    allowed_types = ["image/jpeg", "image/jpg", "image/png"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Only JPG, PNG, and JPEG files are allowed")

    # Create uploads directory if it doesn't exist
    upload_dir = ROOT_DIR / "uploads" / "profile_photos"
    upload_dir.mkdir(parents=True, exist_ok=True)

    # Generate filename with user ID and original extension
    file_extension = file.filename.split('.')[-1].lower()
    filename = f"{current_user['id']}.{file_extension}"
    file_path = upload_dir / filename

    # Save the file
    try:
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to save profile photo")

    return create_response(True, "Profile photo uploaded successfully")

@api_router.get("/auth/profile-photo/{user_id}")
async def get_profile_photo(user_id: str):
    # Check if profile photo exists
    upload_dir = ROOT_DIR / "uploads" / "profile_photos"

    # Try different extensions
    for ext in ["jpg", "jpeg", "png"]:
        file_path = upload_dir / f"{user_id}.{ext}"
        if file_path.exists():
            from fastapi.responses import FileResponse
            return FileResponse(file_path, media_type=f"image/{ext}")

    # If no photo exists, redirect to dicebear avatar
    # Get user name for dicebear seed (this is a simplified approach)
    # In a real app, you might want to cache this or get from DB
    dicebear_url = f"https://api.dicebear.com/7.x/avataaars/svg?seed={user_id}"
    from fastapi.responses import RedirectResponse
    return RedirectResponse(dicebear_url)

# ===== AUTO-ASSIGNMENT HELPER =====
async def auto_assign_complaint(category: str, session: AsyncSession, student_department: str = None):
    """Auto-assign complaint based on category.

    - HOD_CATEGORIES (Academic Issues, Staff Behavior): assigned to the HOD
      of the student's department.
    - All other categories: assigned to staff matching CATEGORY_TO_STAFF_ROLE,
      load-balanced by fewest active complaints.

    Returns (staff_id, staff_name) or (None, None). Never raises.
    """
    try:
        if not category:
            return None, None

        # ── HOD-based assignment ──
        if category in HOD_CATEGORIES:
            if not student_department:
                logger.warning(f"Cannot assign '{category}' complaint — student department is unknown")
                return None, None

            hod_stmt = select(User).where(
                User.role == UserRole.HOD,
                User.department == student_department
            )
            result = await session.execute(hod_stmt)
            hod = result.scalars().first()

            if not hod:
                logger.warning(
                    f"No HOD found for department '{student_department}' "
                    f"to handle '{category}' complaint"
                )
                return None, None

            logger.info(
                f"Auto-assigning '{category}' complaint to HOD "
                f"'{hod.name}' (dept: {student_department})"
            )
            return hod.id, hod.name

        # ── Role-based assignment ──
        target_role = CATEGORY_TO_STAFF_ROLE.get(category)
        if not target_role:
            logger.info(f"No staff role mapping for category '{category}', skipping auto-assign")
            return None, None

        # Find all staff with the matching staff_role
        staff_stmt = select(User).where(
            User.role == UserRole.STAFF,
            User.staff_role == target_role
        )
        result = await session.execute(staff_stmt)
        staff_list = result.scalars().all()

        if not staff_list:
            logger.warning(f"No staff found with role '{target_role}' for category '{category}'")
            return None, None

        # Pick staff with fewest active (non-resolved, non-rejected) complaints
        best_staff = None
        min_count = float('inf')
        for staff in staff_list:
            count_stmt = select(func.count(Complaint.id)).where(
                Complaint.assigned_to == staff.id,
                Complaint.status.notin_([ComplaintStatus.RESOLVED, ComplaintStatus.REJECTED])
            )
            count_result = await session.execute(count_stmt)
            active_count = count_result.scalar() or 0
            if active_count < min_count:
                min_count = active_count
                best_staff = staff

        if best_staff:
            logger.info(f"Auto-assigning to staff '{best_staff.name}' (role: {target_role}, active: {min_count})")
            return best_staff.id, best_staff.name

        return None, None
    except Exception as e:
        logger.error(f"Error in auto_assign_complaint for category '{category}': {str(e)}")
        return None, None

# ===== COMPLAINT ENDPOINTS =====
@api_router.post("/complaints", response_model=ComplaintResponse, status_code=201)
async def create_complaint(complaint_data: ComplaintCreate, current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    # Check for duplicates
    duplicate_id = await check_duplicate_complaint(
        complaint_data.title,
        complaint_data.description,
        current_user["id"],
        session
    )

    if duplicate_id:
        logger.warning(f"Duplicate complaint attempt by user {current_user['id']}. Existing ID: {duplicate_id}")
        raise HTTPException(
            status_code=400,
            detail=f"Similar complaint already exists: {duplicate_id}"
        )

    # Check for foul/offensive language before processing
    from utils.profanity_filter import contains_profanity
    full_text_check = f"{complaint_data.title} {complaint_data.description}"
    if complaint_data.voice_text:
        full_text_check += f" {complaint_data.voice_text}"
    
    if contains_profanity(full_text_check):
        raise HTTPException(
            status_code=400,
            detail="Unwanted or offensive language detected. Please do not send such messages."
        )

    # AI Analysis
    full_text = f"{complaint_data.title}. {complaint_data.description}"
    if complaint_data.voice_text:
        full_text += f" {complaint_data.voice_text}"

    analysis = await analyze_complaint_with_ai(full_text)

    # Determine final category: prefer client-provided category if present, else AI analysis
    final_category = None
    
    # Check if category is provided and not empty/whitespace
    if complaint_data.category and complaint_data.category.strip():
        final_category = complaint_data.category
    
    # Only fall back to AI if no user category provided
    if not final_category:
        final_category = analysis.category or "Other"

    # Look up student's department for HOD-based assignment
    student_user = await session.get(User, current_user["id"])
    student_dept = student_user.department if student_user else current_user.get("department")

    # Create complaint
    complaint_obj = Complaint(
        title=complaint_data.title,
        description=complaint_data.description,
        voice_text=complaint_data.voice_text,
        status=ComplaintStatus.SUBMITTED,
        category=final_category,
        priority=analysis.priority,
        sentiment=analysis.sentiment,
        foul_language_detected=analysis.foul_language_detected,
        foul_language_severity=analysis.foul_language_severity,
        is_anonymous=complaint_data.is_anonymous,
        student_id=current_user["id"],
        student_name=None if complaint_data.is_anonymous else current_user["name"],
        student_email=current_user["email"],
        student_department=student_dept,
        support_count=0,
        supported_by=[],
        responses=[],
        timeline=[
            {
                "status": ComplaintStatus.SUBMITTED,
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "note": "Complaint submitted"
            }
        ]
    )

    # Auto-assign complaint based on category → staff role / HOD mapping
    try:
        assigned_staff_id, assigned_staff_name = await auto_assign_complaint(
            final_category, session, student_department=student_dept
        )
        if assigned_staff_id:
            complaint_obj.assigned_to = assigned_staff_id
            complaint_obj.assigned_to_name = assigned_staff_name
            complaint_obj.assigned_at = datetime.now(timezone.utc)
            complaint_obj.timeline.append({
                "status": "auto_assigned",
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "note": f"Auto-assigned to {assigned_staff_name}"
            })
    except Exception as e:
        logger.error(f"Auto-assignment failed for category '{final_category}': {str(e)}")
        # Continue without assignment — complaint is still created successfully

    session.add(complaint_obj)
    
    # Debug logging before commit
    logger.info(f"Creating complaint - student_id: {complaint_obj.student_id}, student_name: {complaint_obj.student_name}, student_email: {complaint_obj.student_email}")
    
    await session.commit()
    await session.refresh(complaint_obj)
    
    # Debug logging after commit
    logger.info(f"Complaint created - ID: {complaint_obj.id}, student_id saved: {complaint_obj.student_id}, assigned_to: {complaint_obj.assigned_to}")

    response_data = {
        "id": complaint_obj.id,
        "title": complaint_obj.title,
        "description": complaint_obj.description,
        "status": complaint_obj.status,
        "category": complaint_obj.category,
        "priority": complaint_obj.priority,
        "sentiment": complaint_obj.sentiment,
        "foul_language_severity": complaint_obj.foul_language_severity,
        "foul_language_detected": complaint_obj.foul_language_detected,
        "is_anonymous": complaint_obj.is_anonymous,
        "student_id": complaint_obj.student_id,
        "student_name": complaint_obj.student_name,
        "student_email": complaint_obj.student_email,
        "support_count": complaint_obj.support_count,
        "assigned_to": complaint_obj.assigned_to,
        "assigned_to_name": complaint_obj.assigned_to_name,
        "assigned_at": complaint_obj.assigned_at.isoformat() if complaint_obj.assigned_at else None,
        "created_at": complaint_obj.created_at.isoformat() if complaint_obj.created_at else None,
        "updated_at": complaint_obj.updated_at.isoformat() if complaint_obj.updated_at else None,
        "responses": complaint_obj.responses,
        "timeline": complaint_obj.timeline
    }

    return ComplaintResponse(**response_data)

@api_router.post("/complaints/transcribe")
async def transcribe_voice(audio_base64: str, current_user: dict = Depends(get_current_user)):
    text = await transcribe_audio_with_whisper(audio_base64)
    return {"text": text}

@api_router.get("/complaints")
async def get_complaints(current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    stmt = select(Complaint)

    # Role-based filtering
    if current_user["role"] == UserRole.STUDENT:
        stmt = stmt.where(Complaint.student_id == current_user["id"])
    elif current_user["role"] == UserRole.STAFF:
        # Staff should only see complaints assigned to them
        stmt = stmt.where(Complaint.assigned_to == current_user["id"])
    elif current_user["role"] == UserRole.HOD:
        # HOD only sees HOD-category complaints (Academic Issues / Staff Behavior)
        # from their department or assigned to them
        stmt = stmt.where(
            Complaint.category.in_(list(HOD_CATEGORIES)),
            or_(
                Complaint.assigned_to == current_user["id"],
                Complaint.student_department == current_user.get("department")
            )
        )
    # Admin, Principal see all complaints

    stmt = stmt.order_by(Complaint.created_at.desc()).limit(1000)
    res = await session.execute(stmt)
    complaints = res.scalars().all()

    def _to_response(c: Complaint):
        return {
            "id": c.id,
            "title": c.title,
            "description": c.description,
            "status": c.status,
            "category": c.category,
            "priority": c.priority,
            "sentiment": c.sentiment,
            "foul_language_severity": c.foul_language_severity,
            "foul_language_detected": c.foul_language_detected,
            "is_anonymous": c.is_anonymous,
            "student_id": c.student_id,
            "student_name": c.student_name,
            "support_count": c.support_count,
            "assigned_to": c.assigned_to,
            "assigned_to_name": c.assigned_to_name,
            "assigned_at": c.assigned_at.isoformat() if c.assigned_at else None,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
            "responses": c.responses,
            "timeline": c.timeline
        }

    return create_response(True, "Complaints retrieved successfully", [_to_response(c) for c in complaints])

@api_router.get("/complaints/{complaint_id}", response_model=ComplaintResponse)
async def get_complaint(complaint_id: str, current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    complaint = await session.get(Complaint, complaint_id)
    if not complaint:
        raise HTTPException(status_code=404, detail="Complaint not found")


    return ComplaintResponse(
        id=complaint.id,
        title=complaint.title,
        description=complaint.description,
        status=complaint.status,
        category=complaint.category,
        priority=complaint.priority,
        sentiment=complaint.sentiment,
        foul_language_severity=complaint.foul_language_severity,
        foul_language_detected=complaint.foul_language_detected,
        is_anonymous=complaint.is_anonymous,
        student_id=complaint.student_id,
        student_name=complaint.student_name,
        student_email=complaint.student_email,
        support_count=complaint.support_count,
        assigned_to=complaint.assigned_to,
        assigned_to_name=complaint.assigned_to_name,
        assigned_at=complaint.assigned_at.isoformat() if complaint.assigned_at else None,
        created_at=complaint.created_at.isoformat() if complaint.created_at else None,
        updated_at=complaint.updated_at.isoformat() if complaint.updated_at else None,
        responses=complaint.responses,
        timeline=complaint.timeline
    )

@api_router.put("/complaints/{complaint_id}")
async def update_complaint(complaint_id: str, update_data: ComplaintUpdate, current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    logger.info(f"PUT /complaints/{complaint_id} - User: {current_user['id']}, Data: {update_data.model_dump()}")

    # Only admin, staff, hod, principal can update
    if current_user["role"] not in [UserRole.ADMIN, UserRole.STAFF, UserRole.HOD, UserRole.PRINCIPAL]:
        logger.warning(f"Permission denied for user {current_user['id']} to update complaint {complaint_id}")
        return create_response(False, "Permission denied", status_code=403)

    complaint_obj = await session.get(Complaint, complaint_id)
    if not complaint_obj:
        logger.warning(f"Complaint {complaint_id} not found")
        return create_response(False, "Complaint not found", status_code=404)


    now = datetime.now(timezone.utc).isoformat()

    if update_data.status:
        # HOD cannot change complaint status (no Accept / Reject / Close)
        if current_user["role"] == UserRole.HOD:
            return create_response(
                False,
                "HOD cannot accept, reject, or change complaint status. You can only view and assign complaints.",
                status_code=403
            )
        complaint_obj.status = update_data.status
        timeline = complaint_obj.timeline or []
        timeline.append({
            "status": update_data.status,
            "timestamp": now,
            "note": f"Status updated to {update_data.status}",
            "updated_by": current_user["name"]
        })
        complaint_obj.timeline = timeline

    if update_data.response_text:
        responses = complaint_obj.responses or []
        responses.append({
            "text": update_data.response_text,
            "responder_name": current_user["name"],
            "responder_role": current_user["role"],
            "timestamp": now
        })
        complaint_obj.responses = responses

    if update_data.assigned_to:
        # HOD-only categories: only the HOD of the student's department can reassign
        if complaint_obj.category in HOD_CATEGORIES:
            if current_user["role"] != UserRole.HOD:
                return create_response(
                    False,
                    f"Only the HOD can reassign '{complaint_obj.category}' complaints.",
                    status_code=403
                )
            if current_user.get("department") != complaint_obj.student_department:
                return create_response(
                    False,
                    "You can only reassign complaints from your own department.",
                    status_code=403
                )
        # First, fetch the user to check for conflict of interest
        try:
            user_stmt = select(User).where(User.id == update_data.assigned_to)
            res = await session.execute(user_stmt)
            assigned_user = res.scalars().first()
            
            if not assigned_user:
                return create_response(False, "Staff member not found", status_code=404)

            # For HOD-category complaints, target staff must be in the same department
            if complaint_obj.category in HOD_CATEGORIES:
                if assigned_user.department != complaint_obj.student_department:
                    return create_response(
                        False,
                        f"Cannot assign to {assigned_user.name} — they are not in the '{complaint_obj.student_department}' department.",
                        status_code=400
                    )
                # HOD can only assign to Assistant Professors within the department
                if current_user["role"] == UserRole.HOD:
                    if assigned_user.staff_role != StaffRole.ASSISTANT_PROFESSOR:
                        return create_response(
                            False,
                            f"HOD can only assign Academic/Staff Behaviour complaints to Assistant Professors. "
                            f"'{assigned_user.name}' has role '{assigned_user.staff_role}'.",
                            status_code=400
                        )
            
            # CONFLICT OF INTEREST CHECK: Prevent assigning staff mentioned in the complaint
            from utils.conflict_detection import is_staff_mentioned_in_complaint
            if is_staff_mentioned_in_complaint(
                assigned_user.name,
                complaint_obj.title,
                complaint_obj.description
            ):
                logger.warning(f"Conflict of interest: Cannot assign {assigned_user.name} to complaint {complaint_id} - mentioned in complaint")
                return create_response(
                    False,
                    f"Cannot assign {assigned_user.name} to this complaint - they are mentioned in the complaint. Please select a different staff member.",
                    status_code=400
                )
            
            # Set assigned user id and name
            complaint_obj.assigned_to = update_data.assigned_to
            complaint_obj.assigned_to_name = assigned_user.name
            complaint_obj.assigned_at = datetime.now(timezone.utc)

            # append a timeline entry for the assignment
            timeline = complaint_obj.timeline or []
            timeline.append({
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "action": f"Reassigned to {assigned_user.name}" if complaint_obj.assigned_to else f"Assigned to {assigned_user.name}",
                "by": current_user["name"]
            })
            complaint_obj.timeline = timeline
            # ensure status moves to in_progress when assigned
            complaint_obj.status = ComplaintStatus.IN_PROGRESS
        except Exception as e:
            logger.error(f"Error assigning complaint: {str(e)}")
            return create_response(False, f"Error assigning complaint: {str(e)}", status_code=500)

    # update timestamp
    complaint_obj.updated_at = datetime.now(timezone.utc)

    session.add(complaint_obj)
    await session.commit()
    await session.refresh(complaint_obj)

    logger.info(f"Complaint {complaint_id} updated successfully")

    data = {
        "id": complaint_obj.id,
        "title": complaint_obj.title,
        "description": complaint_obj.description,
        "status": complaint_obj.status,
        "category": complaint_obj.category,
        "priority": complaint_obj.priority,
        "sentiment": complaint_obj.sentiment,
        "foul_language_severity": complaint_obj.foul_language_severity,
        "foul_language_detected": complaint_obj.foul_language_detected,
        "is_anonymous": complaint_obj.is_anonymous,
        "student_id": complaint_obj.student_id,
        "student_name": complaint_obj.student_name,
        "support_count": complaint_obj.support_count,
        "created_at": complaint_obj.created_at.isoformat() if complaint_obj.created_at else None,
        "updated_at": complaint_obj.updated_at.isoformat() if complaint_obj.updated_at else None,
        "responses": complaint_obj.responses,
        "timeline": complaint_obj.timeline
    }
    return create_response(True, "Complaint updated successfully", data)

@api_router.get("/complaints/{complaint_id}/eligible-staff")
async def get_eligible_staff(complaint_id: str, current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    """
    Get list of eligible staff for assignment, filtering out conflicts of interest.
    - For HOD categories: returns only Assistant Professors in student's department
    - For other categories: returns all staff
    - Always filters out staff mentioned in the complaint
    """
    complaint = await session.get(Complaint, complaint_id)
    if not complaint:
        raise HTTPException(status_code=404, detail="Complaint not found")

    # 1. Base query for staff
    stmt = select(User).where(User.role == UserRole.STAFF)
    
    # 2. Apply HOD-category strict filtering
    if complaint.category in HOD_CATEGORIES:
        # Must be in student's department AND Assistant Professor
        stmt = stmt.where(
            User.department == complaint.student_department,
            User.staff_role == StaffRole.ASSISTANT_PROFESSOR
        )
    
    result = await session.execute(stmt)
    all_staff = result.scalars().all()
    
    # 3. Filter out conflicts of interest (mentioned in complaint)
    from utils.conflict_detection import is_staff_mentioned_in_complaint
    
    eligible = []
    excluded_names = []
    
    for staff in all_staff:
        if is_staff_mentioned_in_complaint(staff.name, complaint.title, complaint.description):
            excluded_names.append(staff.name)
        else:
            eligible.append({
                "id": staff.id,
                "name": staff.name,
                "department": staff.department,
                "staff_role": staff.staff_role
            })
            
    return {
        "staff": eligible,
        "excluded_count": len(excluded_names),
        "excluded_names": excluded_names
    }

@api_router.delete("/complaints/{complaint_id}")
async def delete_complaint(complaint_id: str, confirm: bool = False, current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    """
    Delete a complaint (Complete action).
    
    Rules:
    - Admin can delete any complaint
    - Staff can only delete complaints assigned to them AND with status 'resolved'
    - Requires confirm=true to prevent accidental deletion
    """
    logger.info(f"DELETE /complaints/{complaint_id} - User: {current_user['id']}, Role: {current_user['role']}, Confirm: {confirm}")

    complaint_obj = await session.get(Complaint, complaint_id)
    if not complaint_obj:
        logger.warning(f"Complaint {complaint_id} not found")
        return create_response(False, "Complaint not found", status_code=404)

    # Check confirmation parameter
    if not confirm:
        return create_response(False, "Confirmation required. Pass confirm=true to delete.", status_code=400)

    user_role = current_user["role"]
    user_id = current_user["id"]

    # Admin can delete any complaint
    if user_role == UserRole.ADMIN:
        await session.delete(complaint_obj)
        await session.commit()
        logger.info(f"Complaint {complaint_id} deleted by Admin {user_id}")
        return create_response(True, "Complaint deleted successfully")

    # Staff can only delete complaints assigned to them with status 'resolved'
    if user_role == UserRole.STAFF:
        if complaint_obj.assigned_to != user_id:
            logger.warning(f"Staff {user_id} tried to delete complaint {complaint_id} not assigned to them")
            return create_response(False, "You can only complete complaints assigned to you", status_code=403)
        
        if complaint_obj.status != ComplaintStatus.RESOLVED:
            logger.warning(f"Staff {user_id} tried to delete complaint {complaint_id} with status {complaint_obj.status}")
            return create_response(False, "Complaint must be resolved before completing", status_code=400)
        
        await session.delete(complaint_obj)
        await session.commit()
        logger.info(f"Complaint {complaint_id} completed and deleted by Staff {user_id}")
        return create_response(True, "Complaint completed successfully")

    # Other roles cannot delete
    logger.warning(f"Permission denied for user {current_user['id']} (role: {user_role}) to delete complaint {complaint_id}")
    return create_response(False, "Permission denied", status_code=403)

@api_router.post("/complaints/{complaint_id}/support")
async def support_complaint(complaint_id: str, current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    complaint_obj = await session.get(Complaint, complaint_id)
    if not complaint_obj:
        return create_response(False, "Complaint not found", status_code=404)

    supported_by = complaint_obj.supported_by or []

    if current_user["id"] in supported_by:
        # Remove support
        supported_by = [uid for uid in supported_by if uid != current_user["id"]]
        support_count = max(0, complaint_obj.support_count - 1)
    else:
        # Add support
        supported_by.append(current_user["id"])
        support_count = complaint_obj.support_count + 1

    complaint_obj.supported_by = supported_by
    complaint_obj.support_count = support_count

    session.add(complaint_obj)
    await session.commit()
    await session.refresh(complaint_obj)

    data = {"support_count": support_count, "user_supported": current_user["id"] in supported_by}
    return create_response(True, "Support updated successfully", data)

@api_router.post("/complaints/{complaint_id}/status")
async def update_complaint_status(complaint_id: str, status: str, remarks: str, user=Depends(require_roles("HOD", "PRINCIPAL", "ADMIN")), session: AsyncSession = Depends(get_session)):
    # Only HOD/Principal/Admin reach here
    # Log status change with user['role'], timestamp, remarks (use existing fields)
    complaint_obj = await session.get(Complaint, complaint_id)
    if not complaint_obj:
        return create_response(False, "Complaint not found", status_code=404)

    # CONFLICT OF INTEREST CHECK: Prevent user from verifying complaints that mention them
    from utils.conflict_detection import can_user_verify_complaint, get_escalation_authority
    can_verify, reason = can_user_verify_complaint(
        user["name"],
        user["role"],
        complaint_obj.title,
        complaint_obj.description
    )
    
    if not can_verify:
        escalate_to = get_escalation_authority(user["role"])
        logger.warning(f"Conflict of interest: {user['name']} ({user['role']}) cannot verify complaint {complaint_id} - mentioned in complaint")
        return create_response(
            False,
            reason,
            {"escalate_to": escalate_to} if escalate_to else None,
            status_code=403
        )

    # Update status and log
    complaint_obj.status = status
    complaint_obj.timeline.append({
        "status": status,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "note": remarks,
        "updated_by": user["name"]
    })

    session.add(complaint_obj)
    await session.commit()
    await session.refresh(complaint_obj)

    return {"success": True}

@api_router.get("/complaints/{complaint_id}/eligible-staff")
async def get_eligible_staff(
    complaint_id: str,
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """
    Get list of staff members eligible to handle this complaint.
    - For HOD categories (Academic Issues, Staff Behavior): returns only staff in the
      student's department (only HOD of that department may call this).
    - For all other categories: returns all staff, excluding those mentioned in the complaint.
    """
    # Only HOD/Principal/Admin can access this
    if current_user["role"] not in [UserRole.HOD, UserRole.PRINCIPAL, UserRole.ADMIN]:
        return create_response(False, "Permission denied", status_code=403)
    
    complaint = await session.get(Complaint, complaint_id)
    if not complaint:
        return create_response(False, "Complaint not found", status_code=404)

    # HOD-category complaints: only departmental HOD can fetch, scoped to department
    is_hod_category = complaint.category in HOD_CATEGORIES
    if is_hod_category:
        if current_user["role"] != UserRole.HOD:
            return create_response(
                False,
                f"Only the HOD can reassign '{complaint.category}' complaints.",
                status_code=403
            )
        if current_user.get("department") != complaint.student_department:
            return create_response(
                False,
                "You can only view eligible staff for complaints from your department.",
                status_code=403
            )
        # Fetch staff in the same department (staff + hod roles)
        stmt = select(User).where(
            User.role.in_([UserRole.STAFF, UserRole.HOD]),
            User.department == complaint.student_department
        )
    else:
        # All staff for non-HOD categories
        stmt = select(User).where(User.role == UserRole.STAFF)
    
    res = await session.execute(stmt)
    all_staff = res.scalars().all()
    
    # Filter out mentioned staff using conflict detection
    from utils.conflict_detection import get_eligible_staff_for_assignment
    eligible_staff, excluded_staff = get_eligible_staff_for_assignment(complaint, all_staff)
    
    eligible_list = [
        {"id": s.id, "name": s.name, "department": s.department, "staff_role": s.staff_role}
        for s in eligible_staff
    ]
    
    excluded_names = [s.name for s in excluded_staff]
    
    logger.info(f"Eligible staff for complaint {complaint_id}: {len(eligible_list)} eligible, {len(excluded_names)} excluded")
    
    return create_response(
        True,
        "Eligible staff retrieved",
        {
            "staff": eligible_list,
            "excluded_count": len(excluded_staff),
            "excluded_names": excluded_names,
            "is_hod_category": is_hod_category
        }
    )

# ===== ANALYTICS ENDPOINTS =====
@api_router.get("/analytics/overview")
async def get_analytics_overview(current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    # Only admin, principal, hod can access
    if current_user["role"] not in [UserRole.ADMIN, UserRole.PRINCIPAL, UserRole.HOD]:
        raise HTTPException(status_code=403, detail="Permission denied")

    total_stmt = select(func.count()).select_from(Complaint)
    resolved_stmt = select(func.count()).where(Complaint.status == ComplaintStatus.RESOLVED)
    pending_stmt = select(func.count()).where(Complaint.status.in_([ComplaintStatus.SUBMITTED, ComplaintStatus.REVIEWED, ComplaintStatus.IN_PROGRESS]))

    total = (await session.execute(total_stmt)).scalar() or 0
    resolved = (await session.execute(resolved_stmt)).scalar() or 0
    pending = (await session.execute(pending_stmt)).scalar() or 0

    # Calculate average resolution time for resolved complaints
    avg_resolution_time = 0.0
    if resolved > 0:
        resolved_complaints_stmt = select(Complaint.created_at, Complaint.updated_at).where(
            Complaint.status == ComplaintStatus.RESOLVED,
            Complaint.created_at.isnot(None),
            Complaint.updated_at.isnot(None)
        )
        resolved_complaints = (await session.execute(resolved_complaints_stmt)).all()
        
        if resolved_complaints:
            total_resolution_days = 0.0
            for created_at, updated_at in resolved_complaints:
                resolution_time = (updated_at - created_at).total_seconds() / 86400  # Convert to days
                total_resolution_days += resolution_time
            avg_resolution_time = round(total_resolution_days / len(resolved_complaints), 1)

    # Calculate satisfaction rate based on positive sentiment
    satisfaction_rate = 0.0
    if total > 0:
        positive_stmt = select(func.count()).where(Complaint.sentiment == SentimentType.POSITIVE)
        positive_count = (await session.execute(positive_stmt)).scalar() or 0
        satisfaction_rate = round((positive_count / total * 100), 0)

    # By category
    cat_stmt = select(Complaint.category, func.count()).group_by(Complaint.category)
    prio_stmt = select(Complaint.priority, func.count()).group_by(Complaint.priority)
    sent_stmt = select(Complaint.sentiment, func.count()).group_by(Complaint.sentiment)

    cat_rows = (await session.execute(cat_stmt)).all()
    prio_rows = (await session.execute(prio_stmt)).all()
    sent_rows = (await session.execute(sent_stmt)).all()

    data = {
        "total_complaints": total,
        "resolved_complaints": resolved,
        "pending_complaints": pending,
        "avg_resolution_time": avg_resolution_time,
        "satisfaction_rate": satisfaction_rate,
        "resolution_rate": round((resolved / total * 100) if total > 0 else 0, 2),
        "by_category": {k: v for k, v in cat_rows},
        "by_priority": {k: v for k, v in prio_rows},
        "by_sentiment": {k: v for k, v in sent_rows}
    }
    return create_response(True, "Analytics retrieved successfully", data)

@api_router.get("/analytics/staff-performance")
async def get_staff_performance(current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.PRINCIPAL, UserRole.HOD]:
        return create_response(False, "Permission denied", status_code=403)

    stmt = select(User).where(User.role == UserRole.STAFF)
    res = await session.execute(stmt)
    staff_users = res.scalars().all()

    performance = []
    for staff in staff_users:
        assigned = (await session.execute(select(func.count()).where(Complaint.assigned_to == staff.id))).scalar() or 0
        resolved = (await session.execute(select(func.count()).where(Complaint.assigned_to == staff.id, Complaint.status == ComplaintStatus.RESOLVED))).scalar() or 0
        pending = assigned - resolved

        performance.append({
            "staff_id": staff.id,
            "staff_name": staff.name,
            "total_complaints": assigned,
            "resolved_complaints": resolved,
            "pending_complaints": pending,
            "resolution_rate": round((resolved / assigned * 100) if assigned > 0 else 0, 2)
        })

    return create_response(True, "Staff performance retrieved successfully", performance)

# ===== STAFF SELF-SERVICE ENDPOINTS =====

@api_router.get("/staff/my-performance")
async def get_my_performance(current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    """Get performance statistics for the logged-in staff member"""
    if current_user["role"] != UserRole.STAFF:
        return create_response(False, "Permission denied. This endpoint is for staff only.", status_code=403)
    
    staff_id = current_user["id"]
    
    # Get counts by status
    total_stmt = select(func.count()).where(Complaint.assigned_to == staff_id)
    resolved_stmt = select(func.count()).where(
        Complaint.assigned_to == staff_id,
        Complaint.status == ComplaintStatus.RESOLVED
    )
    rejected_stmt = select(func.count()).where(
        Complaint.assigned_to == staff_id,
        Complaint.status == ComplaintStatus.REJECTED
    )
    in_progress_stmt = select(func.count()).where(
        Complaint.assigned_to == staff_id,
        Complaint.status == ComplaintStatus.IN_PROGRESS
    )
    submitted_stmt = select(func.count()).where(
        Complaint.assigned_to == staff_id,
        Complaint.status == ComplaintStatus.SUBMITTED
    )
    reviewed_stmt = select(func.count()).where(
        Complaint.assigned_to == staff_id,
        Complaint.status == ComplaintStatus.REVIEWED
    )
    
    total = (await session.execute(total_stmt)).scalar() or 0
    resolved = (await session.execute(resolved_stmt)).scalar() or 0
    rejected = (await session.execute(rejected_stmt)).scalar() or 0
    in_progress = (await session.execute(in_progress_stmt)).scalar() or 0
    submitted = (await session.execute(submitted_stmt)).scalar() or 0
    reviewed = (await session.execute(reviewed_stmt)).scalar() or 0
    
    pending = submitted + reviewed + in_progress
    
    # Calculate average resolution time for resolved complaints
    resolved_complaints_stmt = select(Complaint).where(
        Complaint.assigned_to == staff_id,
        Complaint.status == ComplaintStatus.RESOLVED
    )
    res = await session.execute(resolved_complaints_stmt)
    resolved_complaints = res.scalars().all()
    
    avg_resolution_days = 0.0
    if resolved_complaints:
        total_days = 0.0
        count = 0
        for c in resolved_complaints:
            if c.updated_at and c.assigned_at:
                diff = c.updated_at - c.assigned_at
                total_days += diff.total_seconds() / 86400  # Convert to days
                count += 1
        if count > 0:
            avg_resolution_days = round(total_days / count, 1)
    
    data = {
        "total_assigned": total,
        "resolved": resolved,
        "rejected": rejected,
        "pending": pending,
        "in_progress": in_progress,
        "resolution_rate": round((resolved / total * 100) if total > 0 else 0, 1),
        "avg_resolution_time_days": avg_resolution_days,
        "staff_name": current_user["name"]
    }
    
    return create_response(True, "Performance retrieved successfully", data)


@api_router.get("/staff/my-complaints")
async def get_my_complaints(current_user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_session)):
    """Get detailed list of complaints assigned to the logged-in staff member"""
    if current_user["role"] != UserRole.STAFF:
        return create_response(False, "Permission denied. This endpoint is for staff only.", status_code=403)
    
    staff_id = current_user["id"]
    
    stmt = select(Complaint).where(Complaint.assigned_to == staff_id).order_by(Complaint.created_at.desc())
    res = await session.execute(stmt)
    complaints = res.scalars().all()
    
    def _to_response(c: Complaint):
        # Get resolution outcome from last response or status
        resolution_outcome = None
        if c.status in [ComplaintStatus.RESOLVED, ComplaintStatus.REJECTED]:
            resolution_outcome = c.status.replace("_", " ").title()
            if c.responses and len(c.responses) > 0:
                last_response = c.responses[-1]
                if isinstance(last_response, dict) and "text" in last_response:
                    resolution_outcome = last_response["text"][:100] + ("..." if len(last_response.get("text", "")) > 100 else "")
        
        return {
            "id": c.id,
            "title": c.title,
            "description": c.description[:100] + "..." if len(c.description) > 100 else c.description,
            "category": c.category,
            "status": c.status,
            "priority": c.priority,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
            "assigned_at": c.assigned_at.isoformat() if c.assigned_at else None,
            "resolution_outcome": resolution_outcome,
            "student_name": c.student_name if not c.is_anonymous else "Anonymous"
        }
    
    return create_response(True, "Complaints retrieved successfully", [_to_response(c) for c in complaints])


@api_router.get("/staff/report/export")
async def export_staff_report(
    format: str = "excel",
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Export staff performance report as PDF or Excel"""
    if current_user["role"] != UserRole.STAFF:
        return create_response(False, "Permission denied. This endpoint is for staff only.", status_code=403)
    
    staff_id = current_user["id"]
    staff_name = current_user["name"]
    
    # Get complaints
    stmt = select(Complaint).where(Complaint.assigned_to == staff_id).order_by(Complaint.created_at.desc())
    res = await session.execute(stmt)
    complaints = res.scalars().all()
    
    # Get stats
    total = len(complaints)
    resolved = sum(1 for c in complaints if c.status == ComplaintStatus.RESOLVED)
    rejected = sum(1 for c in complaints if c.status == ComplaintStatus.REJECTED)
    pending = total - resolved - rejected
    
    if format.lower() == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from fastapi.responses import Response
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Performance Report"
            
            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = f"Performance Report - {staff_name}"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Summary section
            ws['A3'] = "Summary Statistics"
            ws['A3'].font = Font(bold=True, size=12)
            ws['A4'] = "Total Assigned:"
            ws['B4'] = total
            ws['A5'] = "Resolved:"
            ws['B5'] = resolved
            ws['A6'] = "Rejected:"
            ws['B6'] = rejected
            ws['A7'] = "Pending:"
            ws['B7'] = pending
            ws['A8'] = "Resolution Rate:"
            ws['B8'] = f"{round((resolved / total * 100) if total > 0 else 0, 1)}%"
            
            # Complaints table header
            headers = ["Complaint ID", "Title", "Category", "Status", "Priority", "Date", "Resolution"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=10, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            # Complaints data
            for row_idx, c in enumerate(complaints, 11):
                resolution = ""
                if c.status in [ComplaintStatus.RESOLVED, ComplaintStatus.REJECTED]:
                    resolution = c.status.replace("_", " ").title()
                
                data = [
                    c.id[:8] + "...",
                    c.title[:30] + ("..." if len(c.title) > 30 else ""),
                    c.category or "N/A",
                    c.status.replace("_", " ").title(),
                    c.priority or "N/A",
                    c.created_at.strftime("%Y-%m-%d") if c.created_at else "N/A",
                    resolution
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=performance_report_{staff_id[:8]}.xlsx"
                }
            )
        except ImportError:
            return create_response(False, "Excel export not available. Please install openpyxl.", status_code=500)
    
    elif format.lower() == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from fastapi.responses import Response
            
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(letter))
            styles = getSampleStyleSheet()
            elements = []
            
            # Title
            title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1)
            elements.append(Paragraph(f"Performance Report - {staff_name}", title_style))
            elements.append(Spacer(1, 20))
            
            # Summary
            summary_data = [
                ["Total Assigned", str(total)],
                ["Resolved", str(resolved)],
                ["Rejected", str(rejected)],
                ["Pending", str(pending)],
                ["Resolution Rate", f"{round((resolved / total * 100) if total > 0 else 0, 1)}%"]
            ]
            summary_table = Table(summary_data, colWidths=[150, 100])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 30))
            
            # Complaints table
            elements.append(Paragraph("Complaint Details", styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            table_data = [["ID", "Title", "Category", "Status", "Priority", "Date"]]
            for c in complaints[:50]:  # Limit to 50 for PDF
                table_data.append([
                    c.id[:8] + "...",
                    c.title[:25] + ("..." if len(c.title) > 25 else ""),
                    c.category or "N/A",
                    c.status.replace("_", " ").title(),
                    c.priority or "N/A",
                    c.created_at.strftime("%Y-%m-%d") if c.created_at else "N/A"
                ])
            
            complaints_table = Table(table_data, colWidths=[80, 180, 100, 100, 80, 80])
            complaints_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            elements.append(complaints_table)
            
            doc.build(elements)
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename=performance_report_{staff_id[:8]}.pdf"
                }
            )
        except ImportError:
            return create_response(False, "PDF export not available. Please install reportlab.", status_code=500)
    
    else:
        return create_response(False, "Invalid format. Use 'pdf' or 'excel'.", status_code=400)


# ===== STAFF RATING ENDPOINTS =====

def get_current_week_info():
    """Get current ISO week number, year, and week date range."""
    now = datetime.now(timezone.utc)
    week_number = now.isocalendar()[1]
    year = now.isocalendar()[0]
    # Calculate week start (Monday) and end (Sunday)
    week_start = now - timedelta(days=now.weekday())
    week_end = week_start + timedelta(days=6)
    return week_number, year, week_start, week_end


@api_router.post("/ratings")
async def submit_staff_rating(
    rating_data: StaffRatingCreate,
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Submit a staff performance rating (Student only, once per staff per week)."""
    # Only students can submit ratings
    if current_user["role"] != UserRole.STUDENT:
        return create_response(False, "Only students can submit staff ratings", status_code=403)
    
    # Verify staff exists
    staff = await session.get(User, rating_data.staff_id)
    if not staff or staff.role != UserRole.STAFF:
        return create_response(False, "Staff member not found", status_code=404)
    
    # Get current week info
    week_number, year, _, _ = get_current_week_info()
    
    # Check for duplicate rating this week
    stmt = select(StaffRating).where(
        StaffRating.student_id == current_user["id"],
        StaffRating.staff_id == rating_data.staff_id,
        StaffRating.week_number == week_number,
        StaffRating.year == year
    )
    result = await session.execute(stmt)
    existing = result.scalars().first()
    
    if existing:
        return create_response(
            False, 
            f"You have already rated this staff member this week. You can rate them again next week.",
            status_code=400
        )
    
    # Create rating
    new_rating = StaffRating(
        student_id=current_user["id"],
        staff_id=rating_data.staff_id,
        week_number=week_number,
        year=year,
        subject_knowledge=rating_data.subject_knowledge,
        teaching_clarity=rating_data.teaching_clarity,
        student_interaction=rating_data.student_interaction,
        punctuality=rating_data.punctuality,
        overall_effectiveness=rating_data.overall_effectiveness
    )
    
    session.add(new_rating)
    await session.commit()
    await session.refresh(new_rating)
    
    logger.info(f"Staff rating submitted: student={current_user['id']}, staff={rating_data.staff_id}, week={week_number}")
    
    avg_rating = (
        new_rating.subject_knowledge + 
        new_rating.teaching_clarity + 
        new_rating.student_interaction + 
        new_rating.punctuality + 
        new_rating.overall_effectiveness
    ) / 5.0
    
    return create_response(True, "Rating submitted successfully", {
        "id": new_rating.id,
        "staff_name": staff.name,
        "week_number": week_number,
        "year": year,
        "average_rating": round(avg_rating, 2)
    })


@api_router.get("/ratings/staff-list")
async def get_staff_for_rating(
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Get list of registered staff members available for rating."""
    # Only students can access this
    if current_user["role"] != UserRole.STUDENT:
        return create_response(False, "Only students can access staff list for rating", status_code=403)
    
    # Get all staff members
    stmt = select(User).where(User.role == UserRole.STAFF)
    result = await session.execute(stmt)
    staff_list = result.scalars().all()
    
    # Get current week's ratings by this student
    week_number, year, _, _ = get_current_week_info()
    rated_stmt = select(StaffRating.staff_id).where(
        StaffRating.student_id == current_user["id"],
        StaffRating.week_number == week_number,
        StaffRating.year == year
    )
    rated_result = await session.execute(rated_stmt)
    rated_staff_ids = set(r[0] for r in rated_result.all())
    
    staff_data = []
    for staff in staff_list:
        staff_data.append({
            "id": staff.id,
            "name": staff.name,
            "department": staff.department,
            "already_rated_this_week": staff.id in rated_staff_ids
        })
    
    return create_response(True, "Staff list retrieved successfully", {
        "staff": staff_data,
        "week_number": week_number,
        "year": year
    })


@api_router.get("/ratings/my-ratings")
async def get_my_staff_ratings(
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Get current student's submitted ratings for this week."""
    if current_user["role"] != UserRole.STUDENT:
        return create_response(False, "Only students can view their ratings", status_code=403)
    
    week_number, year, week_start, week_end = get_current_week_info()
    
    # Get this student's ratings for current week, joined with staff names
    stmt = select(StaffRating, User.name.label("staff_name")).join(
        User, StaffRating.staff_id == User.id
    ).where(
        StaffRating.student_id == current_user["id"],
        StaffRating.week_number == week_number,
        StaffRating.year == year
    ).order_by(StaffRating.created_at.desc())
    
    result = await session.execute(stmt)
    ratings = result.all()
    
    ratings_data = []
    for rating, staff_name in ratings:
        avg = (
            rating.subject_knowledge + 
            rating.teaching_clarity + 
            rating.student_interaction + 
            rating.punctuality + 
            rating.overall_effectiveness
        ) / 5.0
        
        ratings_data.append({
            "id": rating.id,
            "staff_id": rating.staff_id,
            "staff_name": staff_name,
            "subject_knowledge": rating.subject_knowledge,
            "teaching_clarity": rating.teaching_clarity,
            "student_interaction": rating.student_interaction,
            "punctuality": rating.punctuality,
            "overall_effectiveness": rating.overall_effectiveness,
            "average_rating": round(avg, 2),
            "created_at": rating.created_at.isoformat() if rating.created_at else None
        })
    
    return create_response(True, "Your ratings retrieved successfully", {
        "ratings": ratings_data,
        "week_number": week_number,
        "year": year,
        "week_start": week_start.strftime("%Y-%m-%d"),
        "week_end": week_end.strftime("%Y-%m-%d")
    })


@api_router.get("/ratings/weekly-report")
async def get_weekly_staff_performance(
    week: Optional[int] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Get weekly staff performance report (HOD/Principal/Admin only)."""
    if current_user["role"] not in [UserRole.HOD, UserRole.PRINCIPAL, UserRole.ADMIN]:
        return create_response(False, "Only HOD, Principal, or Admin can access performance reports", status_code=403)
    
    # Use current week if not specified
    current_week, current_year, week_start, week_end = get_current_week_info()
    target_week = week if week else current_week
    target_year = year if year else current_year
    
    # Calculate week date range for target week
    if week or year:
        # Calculate the first day of the target week
        jan1 = datetime(target_year, 1, 1, tzinfo=timezone.utc)
        # Find the first Monday of the year
        days_to_monday = (7 - jan1.weekday()) % 7
        first_monday = jan1 + timedelta(days=days_to_monday)
        # Calculate target week start
        week_start = first_monday + timedelta(weeks=target_week - 1)
        week_end = week_start + timedelta(days=6)
    
    # Get all ratings for the week
    stmt = select(
        StaffRating.staff_id,
        User.name.label("staff_name"),
        User.department,
        func.avg(
            (StaffRating.subject_knowledge + 
             StaffRating.teaching_clarity + 
             StaffRating.student_interaction + 
             StaffRating.punctuality + 
             StaffRating.overall_effectiveness) / 5.0
        ).label("avg_rating"),
        func.count(StaffRating.id).label("total_ratings")
    ).join(
        User, StaffRating.staff_id == User.id
    ).where(
        StaffRating.week_number == target_week,
        StaffRating.year == target_year
    ).group_by(
        StaffRating.staff_id, User.name, User.department
    ).order_by(
        func.avg(
            (StaffRating.subject_knowledge + 
             StaffRating.teaching_clarity + 
             StaffRating.student_interaction + 
             StaffRating.punctuality + 
             StaffRating.overall_effectiveness) / 5.0
        ).desc()
    )
    
    result = await session.execute(stmt)
    staff_performance = result.all()
    
    # Build response with "Best Staff" badge for top performer
    performance_data = []
    total_ratings = 0
    
    for idx, row in enumerate(staff_performance):
        total_ratings += row.total_ratings
        performance_data.append({
            "staff_id": row.staff_id,
            "staff_name": row.staff_name,
            "department": row.department,
            "average_rating": round(float(row.avg_rating), 2),
            "total_ratings": row.total_ratings,
            "is_best_staff": idx == 0  # First one (highest rating) is best staff
        })
    
    return create_response(True, "Weekly performance report retrieved", {
        "week_number": target_week,
        "year": target_year,
        "week_start": week_start.strftime("%Y-%m-%d"),
        "week_end": week_end.strftime("%Y-%m-%d"),
        "staff_performance": performance_data,
        "total_ratings": total_ratings
    })


@api_router.get("/ratings/weekly-report/pdf")
async def download_weekly_performance_pdf(
    week: Optional[int] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Download weekly staff performance report as PDF (HOD/Principal/Admin only)."""
    from fastapi.responses import Response
    
    if current_user["role"] not in [UserRole.HOD, UserRole.PRINCIPAL, UserRole.ADMIN]:
        return create_response(False, "Only HOD, Principal, or Admin can download reports", status_code=403)
    
    # Use current week if not specified
    current_week, current_year, week_start, week_end = get_current_week_info()
    target_week = week if week else current_week
    target_year = year if year else current_year
    
    # Calculate week date range
    if week or year:
        jan1 = datetime(target_year, 1, 1, tzinfo=timezone.utc)
        days_to_monday = (7 - jan1.weekday()) % 7
        first_monday = jan1 + timedelta(days=days_to_monday)
        week_start = first_monday + timedelta(weeks=target_week - 1)
        week_end = week_start + timedelta(days=6)
    
    # Get performance data
    stmt = select(
        StaffRating.staff_id,
        User.name.label("staff_name"),
        User.department,
        func.avg(
            (StaffRating.subject_knowledge + 
             StaffRating.teaching_clarity + 
             StaffRating.student_interaction + 
             StaffRating.punctuality + 
             StaffRating.overall_effectiveness) / 5.0
        ).label("avg_rating"),
        func.count(StaffRating.id).label("total_ratings")
    ).join(
        User, StaffRating.staff_id == User.id
    ).where(
        StaffRating.week_number == target_week,
        StaffRating.year == target_year
    ).group_by(
        StaffRating.staff_id, User.name, User.department
    ).order_by(
        func.avg(
            (StaffRating.subject_knowledge + 
             StaffRating.teaching_clarity + 
             StaffRating.student_interaction + 
             StaffRating.punctuality + 
             StaffRating.overall_effectiveness) / 5.0
        ).desc()
    )
    
    result = await session.execute(stmt)
    staff_performance = result.all()
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=20, textColor=colors.HexColor('#1a365d'))
        elements.append(Paragraph("Weekly Staff Performance Report", title_style))
        elements.append(Spacer(1, 10))
        
        # Report info
        elements.append(Paragraph(f"<b>Week:</b> {target_week} of {target_year}", styles['Normal']))
        elements.append(Paragraph(f"<b>Period:</b> {week_start.strftime('%B %d, %Y')} - {week_end.strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Paragraph(f"<b>Generated:</b> {datetime.now(timezone.utc).strftime('%B %d, %Y at %H:%M UTC')}", styles['Normal']))
        elements.append(Paragraph(f"<b>Generated by:</b> {current_user['name']} ({current_user['role'].upper()})", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        if staff_performance:
            # Summary
            total = len(staff_performance)
            total_ratings = sum(row.total_ratings for row in staff_performance)
            elements.append(Paragraph(f"<b>Total Staff Rated:</b> {total}", styles['Normal']))
            elements.append(Paragraph(f"<b>Total Ratings Received:</b> {total_ratings}", styles['Normal']))
            elements.append(Spacer(1, 15))
            
            # Best Staff Highlight
            best = staff_performance[0]
            highlight_style = ParagraphStyle('Highlight', parent=styles['Normal'], 
                                            fontSize=12, textColor=colors.HexColor('#065f46'),
                                            backColor=colors.HexColor('#d1fae5'), borderPadding=10)
            elements.append(Paragraph(
                f"⭐ <b>Best Staff of the Week:</b> {best.staff_name} (Avg Rating: {round(float(best.avg_rating), 2)}/5.0)",
                highlight_style
            ))
            elements.append(Spacer(1, 20))
            
            # Performance Table
            elements.append(Paragraph("<b>Staff Performance Rankings</b>", styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            table_data = [["Rank", "Staff Name", "Department", "Avg Rating", "Total Ratings", "Badge"]]
            for idx, row in enumerate(staff_performance):
                badge = "⭐ Best Staff" if idx == 0 else ""
                table_data.append([
                    str(idx + 1),
                    row.staff_name,
                    row.department or "N/A",
                    f"{round(float(row.avg_rating), 2)}/5.0",
                    str(row.total_ratings),
                    badge
                ])
            
            perf_table = Table(table_data, colWidths=[40, 120, 100, 80, 80, 80])
            perf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#d1fae5')),  # Highlight best staff row
            ]))
            elements.append(perf_table)
        else:
            elements.append(Paragraph("No ratings submitted for this week.", styles['Normal']))
        
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("--- End of Report ---", styles['Normal']))
        
        doc.build(elements)
        output.seek(0)
        
        filename = f"staff_performance_week{target_week}_{target_year}.pdf"
        return Response(
            content=output.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        return create_response(False, "PDF export not available. Please install reportlab.", status_code=500)


# ===== HOD SEMESTER EVALUATION ENDPOINTS =====

# Helper: get current semester (1=Odd: Jul-Dec, 2=Even: Jan-Jun)
def get_current_semester():
    from datetime import datetime
    now = datetime.now()
    month = now.month
    if month >= 7:  # Jul-Dec = Odd semester
        return 1, now.year
    else:  # Jan-Jun = Even semester
        return 2, now.year

# Pydantic models for HOD evaluation
class StudentHODRatingCreate(BaseModel):
    hod_id: str
    approachability: int = Field(..., ge=1, le=5)
    academic_support: int = Field(..., ge=1, le=5)
    placement_guidance: int = Field(..., ge=1, le=5)
    internship_support: int = Field(..., ge=1, le=5)
    grievance_handling: int = Field(..., ge=1, le=5)
    event_organization: int = Field(..., ge=1, le=5)
    student_motivation: int = Field(..., ge=1, le=5)
    on_duty_permission: int = Field(..., ge=1, le=5)

class StaffHODRatingCreate(BaseModel):
    hod_id: str
    leadership: int = Field(..., ge=1, le=5)
    workload_fairness: int = Field(..., ge=1, le=5)
    staff_coordination: int = Field(..., ge=1, le=5)
    academic_monitoring: int = Field(..., ge=1, le=5)
    research_encouragement: int = Field(..., ge=1, le=5)
    university_communication: int = Field(..., ge=1, le=5)
    conflict_resolution: int = Field(..., ge=1, le=5)
    discipline_maintenance: int = Field(..., ge=1, le=5)

class HODToggleRequest(BaseModel):
    is_open: bool


@api_router.get("/hod-eval/toggle")
async def get_hod_eval_toggle(
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Get the current HOD report toggle status."""
    semester, year = get_current_semester()
    stmt = select(HODReportToggle).where(
        HODReportToggle.semester == semester,
        HODReportToggle.year == year
    )
    result = await session.execute(stmt)
    toggle = result.scalars().first()
    return create_response(True, "Toggle status retrieved", {
        "is_open": toggle.is_open if toggle else False,
        "semester": semester,
        "year": year
    })


@api_router.post("/hod-eval/toggle")
async def set_hod_eval_toggle(
    data: HODToggleRequest,
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Set the HOD report toggle (Principal only)."""
    if current_user["role"] != UserRole.PRINCIPAL:
        raise HTTPException(status_code=403, detail="Only Principal can control the report toggle")
    
    semester, year = get_current_semester()
    stmt = select(HODReportToggle).where(
        HODReportToggle.semester == semester,
        HODReportToggle.year == year
    )
    result = await session.execute(stmt)
    toggle = result.scalars().first()
    
    if toggle:
        toggle.is_open = data.is_open
        toggle.updated_by = current_user["id"]
    else:
        toggle = HODReportToggle(
            is_open=data.is_open,
            semester=semester,
            year=year,
            updated_by=current_user["id"]
        )
        session.add(toggle)
    
    await session.commit()
    status_text = "opened" if data.is_open else "closed"
    return create_response(True, f"Report submission {status_text} successfully", {
        "is_open": data.is_open,
        "semester": semester,
        "year": year
    })


@api_router.post("/hod-eval/student-rating")
async def submit_student_hod_rating(
    data: StudentHODRatingCreate,
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Submit a student's HOD semester rating."""
    if current_user["role"] != UserRole.STUDENT:
        raise HTTPException(status_code=403, detail="Only students can submit student ratings")
    
    semester, year = get_current_semester()
    
    # Check toggle
    toggle_stmt = select(HODReportToggle).where(
        HODReportToggle.semester == semester,
        HODReportToggle.year == year
    )
    toggle_result = await session.execute(toggle_stmt)
    toggle = toggle_result.scalars().first()
    if not toggle or not toggle.is_open:
        raise HTTPException(status_code=403, detail="Report submission is currently closed by the Principal")
    
    # Verify HOD exists
    hod = await session.get(User, data.hod_id)
    if not hod or hod.role != UserRole.HOD:
        raise HTTPException(status_code=404, detail="HOD not found")
    
    # Check for duplicate
    dup_stmt = select(HODRating).where(
        HODRating.rater_id == current_user["id"],
        HODRating.hod_id == data.hod_id,
        HODRating.semester == semester,
        HODRating.year == year
    )
    dup_result = await session.execute(dup_stmt)
    if dup_result.scalars().first():
        raise HTTPException(status_code=400, detail="You have already submitted a rating for this HOD this semester")
    
    # Calculate average
    values = [data.approachability, data.academic_support, data.placement_guidance,
              data.internship_support, data.grievance_handling, data.event_organization,
              data.student_motivation, data.on_duty_permission]
    avg = round(sum(values) / len(values), 2)
    
    rating = HODRating(
        rater_id=current_user["id"],
        rater_role="student",
        hod_id=data.hod_id,
        semester=semester,
        year=year,
        department=current_user.get("department"),
        approachability=data.approachability,
        academic_support=data.academic_support,
        placement_guidance=data.placement_guidance,
        internship_support=data.internship_support,
        grievance_handling=data.grievance_handling,
        event_organization=data.event_organization,
        student_motivation=data.student_motivation,
        on_duty_permission=data.on_duty_permission,
        average_rating=avg
    )
    session.add(rating)
    await session.commit()
    await session.refresh(rating)
    
    return create_response(True, "Student HOD rating submitted successfully", {
        "id": rating.id,
        "average_rating": avg,
        "semester": semester,
        "year": year
    })


@api_router.post("/hod-eval/staff-rating")
async def submit_staff_hod_rating(
    data: StaffHODRatingCreate,
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Submit a staff member's HOD semester rating."""
    if current_user["role"] != UserRole.STAFF:
        raise HTTPException(status_code=403, detail="Only staff can submit staff ratings")
    
    semester, year = get_current_semester()
    
    # Check toggle
    toggle_stmt = select(HODReportToggle).where(
        HODReportToggle.semester == semester,
        HODReportToggle.year == year
    )
    toggle_result = await session.execute(toggle_stmt)
    toggle = toggle_result.scalars().first()
    if not toggle or not toggle.is_open:
        raise HTTPException(status_code=403, detail="Report submission is currently closed by the Principal")
    
    # Verify HOD exists
    hod = await session.get(User, data.hod_id)
    if not hod or hod.role != UserRole.HOD:
        raise HTTPException(status_code=404, detail="HOD not found")
    
    # Check for duplicate
    dup_stmt = select(HODRating).where(
        HODRating.rater_id == current_user["id"],
        HODRating.hod_id == data.hod_id,
        HODRating.semester == semester,
        HODRating.year == year
    )
    dup_result = await session.execute(dup_stmt)
    if dup_result.scalars().first():
        raise HTTPException(status_code=400, detail="You have already submitted a rating for this HOD this semester")
    
    # Calculate average
    values = [data.leadership, data.workload_fairness, data.staff_coordination,
              data.academic_monitoring, data.research_encouragement, data.university_communication,
              data.conflict_resolution, data.discipline_maintenance]
    avg = round(sum(values) / len(values), 2)
    
    rating = HODRating(
        rater_id=current_user["id"],
        rater_role="staff",
        hod_id=data.hod_id,
        semester=semester,
        year=year,
        department=current_user.get("department"),
        leadership=data.leadership,
        workload_fairness=data.workload_fairness,
        staff_coordination=data.staff_coordination,
        academic_monitoring=data.academic_monitoring,
        research_encouragement=data.research_encouragement,
        university_communication=data.university_communication,
        conflict_resolution=data.conflict_resolution,
        discipline_maintenance=data.discipline_maintenance,
        average_rating=avg
    )
    session.add(rating)
    await session.commit()
    await session.refresh(rating)
    
    return create_response(True, "Staff HOD rating submitted successfully", {
        "id": rating.id,
        "average_rating": avg,
        "semester": semester,
        "year": year
    })


@api_router.get("/hod-eval/my-rating")
async def get_my_hod_rating(
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Check if the current user has already submitted an HOD rating this semester."""
    semester, year = get_current_semester()
    
    stmt = select(HODRating).where(
        HODRating.rater_id == current_user["id"],
        HODRating.semester == semester,
        HODRating.year == year
    )
    result = await session.execute(stmt)
    ratings = result.scalars().all()
    
    rating_data = []
    for r in ratings:
        entry = {
            "id": r.id,
            "hod_id": r.hod_id,
            "rater_role": r.rater_role,
            "average_rating": r.average_rating,
            "semester": r.semester,
            "year": r.year,
            "created_at": r.created_at.isoformat() if r.created_at else None
        }
        if r.rater_role == "student":
            entry.update({
                "approachability": r.approachability,
                "academic_support": r.academic_support,
                "placement_guidance": r.placement_guidance,
                "internship_support": r.internship_support,
                "grievance_handling": r.grievance_handling,
                "event_organization": r.event_organization,
                "student_motivation": r.student_motivation,
                "on_duty_permission": r.on_duty_permission,
            })
        else:
            entry.update({
                "leadership": r.leadership,
                "workload_fairness": r.workload_fairness,
                "staff_coordination": r.staff_coordination,
                "academic_monitoring": r.academic_monitoring,
                "research_encouragement": r.research_encouragement,
                "university_communication": r.university_communication,
                "conflict_resolution": r.conflict_resolution,
                "discipline_maintenance": r.discipline_maintenance,
            })
        rating_data.append(entry)
    
    return create_response(True, "My HOD ratings retrieved", {
        "ratings": rating_data,
        "semester": semester,
        "year": year,
        "has_submitted": len(rating_data) > 0
    })


@api_router.get("/hod-eval/dashboard")
async def get_hod_eval_dashboard(
    semester: Optional[int] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Get HOD performance dashboard data (Principal only)."""
    if current_user["role"] != UserRole.PRINCIPAL:
        raise HTTPException(status_code=403, detail="Only Principal can access the dashboard")
    
    if semester is None or year is None:
        semester, year = get_current_semester()
    
    # Get all HODs
    hod_stmt = select(User).where(User.role == UserRole.HOD)
    hod_result = await session.execute(hod_stmt)
    hods = hod_result.scalars().all()
    
    # Get all ratings for this semester
    ratings_stmt = select(HODRating).where(
        HODRating.semester == semester,
        HODRating.year == year
    )
    ratings_result = await session.execute(ratings_stmt)
    all_ratings = ratings_result.scalars().all()
    
    # Build per-HOD data
    hod_performance = []
    for hod in hods:
        hod_ratings = [r for r in all_ratings if r.hod_id == hod.id]
        student_ratings = [r for r in hod_ratings if r.rater_role == "student"]
        staff_ratings = [r for r in hod_ratings if r.rater_role == "staff"]
        
        avg_student = round(sum(r.average_rating for r in student_ratings) / len(student_ratings), 2) if student_ratings else 0
        avg_staff = round(sum(r.average_rating for r in staff_ratings) / len(staff_ratings), 2) if staff_ratings else 0
        
        # Overall = average of student avg and staff avg (only non-zero)
        non_zero = [v for v in [avg_student, avg_staff] if v > 0]
        overall = round(sum(non_zero) / len(non_zero), 2) if non_zero else 0
        
        # Determine performance category
        if overall >= 4.5:
            category = "Excellent"
        elif overall >= 4.0:
            category = "Very Good"
        elif overall >= 3.0:
            category = "Good"
        else:
            category = "Needs Improvement"
        
        # Good vs Bad counts for graphs
        student_good = len([r for r in student_ratings if r.average_rating >= 4])
        student_bad = len([r for r in student_ratings if r.average_rating < 4])
        staff_good = len([r for r in staff_ratings if r.average_rating >= 4])
        staff_bad = len([r for r in staff_ratings if r.average_rating < 4])
        
        # Per-criteria averages for student ratings
        student_criteria_avg = {}
        if student_ratings:
            for field in ['approachability', 'academic_support', 'placement_guidance', 'internship_support',
                          'grievance_handling', 'event_organization', 'student_motivation', 'on_duty_permission']:
                vals = [getattr(r, field) for r in student_ratings if getattr(r, field) is not None]
                student_criteria_avg[field] = round(sum(vals) / len(vals), 2) if vals else 0
        
        # Per-criteria averages for staff ratings
        staff_criteria_avg = {}
        if staff_ratings:
            for field in ['leadership', 'workload_fairness', 'staff_coordination', 'academic_monitoring',
                          'research_encouragement', 'university_communication', 'conflict_resolution', 'discipline_maintenance']:
                vals = [getattr(r, field) for r in staff_ratings if getattr(r, field) is not None]
                staff_criteria_avg[field] = round(sum(vals) / len(vals), 2) if vals else 0
        
        hod_performance.append({
            "hod_id": hod.id,
            "hod_name": hod.name,
            "department": hod.department,
            "total_student_ratings": len(student_ratings),
            "total_staff_ratings": len(staff_ratings),
            "avg_student_rating": avg_student,
            "avg_staff_rating": avg_staff,
            "overall_rating": overall,
            "performance_category": category,
            "student_good": student_good,
            "student_bad": student_bad,
            "staff_good": staff_good,
            "staff_bad": staff_bad,
            "student_criteria_avg": student_criteria_avg,
            "staff_criteria_avg": staff_criteria_avg,
        })
    
    # Sort by overall rating descending for ranking
    hod_performance.sort(key=lambda x: x["overall_rating"], reverse=True)
    
    # Assign rank and best HOD badge
    for i, h in enumerate(hod_performance):
        h["rank"] = i + 1
        h["is_best_hod"] = (i == 0 and h["overall_rating"] > 0)
    
    return create_response(True, "HOD evaluation dashboard data retrieved", {
        "semester": semester,
        "year": year,
        "hod_performance": hod_performance,
        "total_hods": len(hods),
        "total_ratings": len(all_ratings)
    })


@api_router.get("/hod-eval/report/pdf")
async def download_hod_report_pdf(
    semester: Optional[int] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Download HOD performance report as PDF (Principal only)."""
    if current_user["role"] != UserRole.PRINCIPAL:
        raise HTTPException(status_code=403, detail="Only Principal can download the report")
    
    if semester is None or year is None:
        semester, year = get_current_semester()
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from fastapi.responses import Response
        import io
        
        # Get dashboard data
        hod_stmt = select(User).where(User.role == UserRole.HOD)
        hod_result = await session.execute(hod_stmt)
        hods = hod_result.scalars().all()
        
        ratings_stmt = select(HODRating).where(
            HODRating.semester == semester,
            HODRating.year == year
        )
        ratings_result = await session.execute(ratings_stmt)
        all_ratings = ratings_result.scalars().all()
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4,
                                topMargin=0.5*inch, bottomMargin=0.5*inch,
                                leftMargin=0.75*inch, rightMargin=0.75*inch)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                     fontSize=20, spaceAfter=20, textColor=colors.HexColor('#1a1a2e'))
        elements.append(Paragraph("HOD Semester Performance Report", title_style))
        
        semester_name = "Odd Semester" if semester == 1 else "Even Semester"
        elements.append(Paragraph(f"<b>Semester:</b> {semester_name} {year}", styles['Normal']))
        elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Per HOD
        for hod in hods:
            hod_ratings = [r for r in all_ratings if r.hod_id == hod.id]
            student_ratings = [r for r in hod_ratings if r.rater_role == "student"]
            staff_ratings = [r for r in hod_ratings if r.rater_role == "staff"]
            
            avg_student = round(sum(r.average_rating for r in student_ratings) / len(student_ratings), 2) if student_ratings else 0
            avg_staff = round(sum(r.average_rating for r in staff_ratings) / len(staff_ratings), 2) if staff_ratings else 0
            non_zero = [v for v in [avg_student, avg_staff] if v > 0]
            overall = round(sum(non_zero) / len(non_zero), 2) if non_zero else 0
            
            if overall >= 4.5:
                category = "Excellent"
            elif overall >= 4.0:
                category = "Very Good"
            elif overall >= 3.0:
                category = "Good"
            else:
                category = "Needs Improvement"
            
            hod_style = ParagraphStyle('HODName', parent=styles['Heading2'],
                                       fontSize=14, textColor=colors.HexColor('#2d3436'))
            elements.append(Paragraph(f"{hod.name} — {hod.department or 'N/A'}", hod_style))
            elements.append(Spacer(1, 8))
            
            # Summary table
            summary_data = [
                ["Metric", "Value"],
                ["Student Ratings Count", str(len(student_ratings))],
                ["Staff Ratings Count", str(len(staff_ratings))],
                ["Avg Student Rating", f"{avg_student}/5"],
                ["Avg Staff Rating", f"{avg_staff}/5"],
                ["Overall Rating", f"{overall}/5"],
                ["Performance Category", category],
            ]
            
            t = Table(summary_data, colWidths=[3*inch, 2.5*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6c5ce7')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f3f5')]),
            ]))
            elements.append(t)
            
            # Student criteria breakdown if available
            if student_ratings:
                elements.append(Spacer(1, 10))
                elements.append(Paragraph("Student Rating Breakdown:", styles['Heading4']))
                criteria_labels = [
                    ('approachability', 'Approachability'), ('academic_support', 'Academic Support'),
                    ('placement_guidance', 'Placement Guidance'), ('internship_support', 'Internship Support'),
                    ('grievance_handling', 'Grievance Handling'), ('event_organization', 'Event & Workshop Organization'),
                    ('student_motivation', 'Student Motivation'), ('on_duty_permission', 'On Duty Permission')
                ]
                s_data = [["Criteria", "Average"]]
                for field, label in criteria_labels:
                    vals = [getattr(r, field) for r in student_ratings if getattr(r, field) is not None]
                    avg_v = round(sum(vals)/len(vals), 2) if vals else 0
                    s_data.append([label, f"{avg_v}/5"])
                st = Table(s_data, colWidths=[3.5*inch, 2*inch])
                st.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00b894')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f3f5')]),
                ]))
                elements.append(st)
            
            # Staff criteria breakdown if available
            if staff_ratings:
                elements.append(Spacer(1, 10))
                elements.append(Paragraph("Staff Rating Breakdown:", styles['Heading4']))
                criteria_labels = [
                    ('leadership', 'Leadership & Decision Making'), ('workload_fairness', 'Workload Distribution Fairness'),
                    ('staff_coordination', 'Staff Coordination'), ('academic_monitoring', 'Academic Monitoring'),
                    ('research_encouragement', 'Research & FDP Encouragement'), ('university_communication', 'Communication with University'),
                    ('conflict_resolution', 'Conflict Resolution'), ('discipline_maintenance', 'Discipline Maintenance')
                ]
                sf_data = [["Criteria", "Average"]]
                for field, label in criteria_labels:
                    vals = [getattr(r, field) for r in staff_ratings if getattr(r, field) is not None]
                    avg_v = round(sum(vals)/len(vals), 2) if vals else 0
                    sf_data.append([label, f"{avg_v}/5"])
                sft = Table(sf_data, colWidths=[3.5*inch, 2*inch])
                sft.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0984e3')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f3f5')]),
                ]))
                elements.append(sft)
            
            elements.append(Spacer(1, 25))
        
        # Ranking section
        if hods:
            elements.append(Paragraph("HOD Performance Ranking", styles['Heading2']))
            elements.append(Spacer(1, 8))
            rank_data = [["Rank", "HOD Name", "Department", "Overall Rating", "Category"]]
            
            # Build ranking
            ranked = []
            for hod in hods:
                hod_ratings = [r for r in all_ratings if r.hod_id == hod.id]
                student_ratings = [r for r in hod_ratings if r.rater_role == "student"]
                staff_ratings = [r for r in hod_ratings if r.rater_role == "staff"]
                avg_s = round(sum(r.average_rating for r in student_ratings)/len(student_ratings), 2) if student_ratings else 0
                avg_st = round(sum(r.average_rating for r in staff_ratings)/len(staff_ratings), 2) if staff_ratings else 0
                non_zero = [v for v in [avg_s, avg_st] if v > 0]
                ov = round(sum(non_zero)/len(non_zero), 2) if non_zero else 0
                if ov >= 4.5: cat = "Excellent"
                elif ov >= 4.0: cat = "Very Good"
                elif ov >= 3.0: cat = "Good"
                else: cat = "Needs Improvement"
                ranked.append((hod.name, hod.department or "N/A", ov, cat))
            
            ranked.sort(key=lambda x: x[2], reverse=True)
            for i, (name, dept, ov, cat) in enumerate(ranked):
                badge = " 🏆" if i == 0 and ov > 0 else ""
                rank_data.append([str(i+1), f"{name}{badge}", dept, f"{ov}/5", cat])
            
            rt = Table(rank_data, colWidths=[0.5*inch, 1.8*inch, 1.5*inch, 1*inch, 1.2*inch])
            rt.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e17055')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f3f5')]),
            ]))
            elements.append(rt)
        
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("--- End of Report ---", styles['Normal']))
        
        doc.build(elements)
        output.seek(0)
        
        semester_label = "odd" if semester == 1 else "even"
        filename = f"hod_performance_report_{semester_label}_{year}.pdf"
        return Response(
            content=output.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        return create_response(False, "PDF export not available. Please install reportlab.", status_code=500)


# HOD list for rating forms
@api_router.get("/hod-eval/hods")
async def get_hods_for_rating(
    current_user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_session)
):
    """Get list of HODs available for rating."""
    hod_stmt = select(User).where(User.role == UserRole.HOD)
    hod_result = await session.execute(hod_stmt)
    hods = hod_result.scalars().all()
    
    semester, year = get_current_semester()
    
    hod_list = []
    for hod in hods:
        # Check if current user already rated this HOD
        dup_stmt = select(HODRating).where(
            HODRating.rater_id == current_user["id"],
            HODRating.hod_id == hod.id,
            HODRating.semester == semester,
            HODRating.year == year
        )
        dup_result = await session.execute(dup_stmt)
        already_rated = dup_result.scalars().first() is not None
        
        hod_list.append({
            "id": hod.id,
            "name": hod.name,
            "department": hod.department,
            "already_rated": already_rated
        })
    
    return create_response(True, "HOD list retrieved", {
        "hods": hod_list,
        "semester": semester,
        "year": year
    })


# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[
        "http://localhost:3000",  # Next.js dev server
        "http://127.0.0.1:3000",  # Alternative localhost
        "http://localhost:8000",  # In case backend serves frontend
        os.environ.get('FRONTEND_URL', 'http://localhost:3000'),  # Environment variable
    ] + (os.environ.get('CORS_ORIGINS', '').split(',') if os.environ.get('CORS_ORIGINS') else []),
    allow_methods=["*"],
    allow_headers=["*"],
)

# ===== EXCEPTION HANDLERS =====
@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    return create_response(False, exc.detail, status_code=exc.status_code)

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    return create_response(False, "Validation error", {"details": exc.errors()}, status_code=422)

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled exception: {str(exc)}", exc_info=True)
    return create_response(False, "Internal server error", status_code=500)

@app.on_event("shutdown")
async def shutdown_db_client():
    # Dispose SQLAlchemy engine on shutdown if initialized
    try:
        eng = get_engine()
        if eng is not None:
            await eng.dispose()
    except Exception as exc:
        logger.exception("Error disposing DB engine: %s", exc)

def role_checker(allowed_roles):
    def decorator(func):
        @wraps(func)
        def wrapper(current_user: dict = Depends(get_current_user), *args, **kwargs):
            if not current_user:
                raise HTTPException(status_code=401, detail="Not authenticated")
            if current_user.get("role") not in allowed_roles:
                raise HTTPException(status_code=403, detail="You do not have enough permissions to access this resource")
            return func(current_user=current_user, *args, **kwargs)
        return wrapper
    return decorator